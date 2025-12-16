"""
Excel Generator для генерации заполненной Карты поиска
Поддерживает все 36 шагов онбординга и заполняет 9 листов Excel
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from typing import Dict, List
import json
from io import BytesIO

# Assuming these imports exist
# from app.database.models import OnboardingSubmission


class SearchMapGenerator:
    """Генератор заполненной Карты поиска для 36 шагов онбординга"""
    
    TEMPLATE_PATH = "Карта_поиска_для_проектной_работы.xlsx"
    
    def __init__(self, user_submissions: List):
        """
        Args:
            user_submissions: Список OnboardingSubmission для пользователя
        """
        # Filter out submissions where step is None to prevent AttributeError
        self.submissions = {s.step.order: s for s in user_submissions if s.step is not None}
        self.wb = load_workbook(self.TEMPLATE_PATH)
    
    def generate(self) -> bytes:
        """Генерирует полностью заполненный Excel-файл"""
        # День 1
        self._fill_plan_sheet()            # Шаг 3
        self._fill_assessment_sheet()      # Шаг 6
        self._fill_vacancy_sheet()         # Шаг 12
        
        # День 2
        self._fill_active_search_sheet()   # Шаг 19
        self._fill_passive_search_sheet()  # Шаг 22
        self._fill_market_analysis_sheet() # Шаг 25
        
        # День 3
        self._fill_speech_modules_sheet()  # Шаги 28-29
        self._fill_objections_sheet()      # Шаг 31
        
        # Отчет по всем шагам
        self._add_onboarding_report_sheet() # Шаги 1-36
        
        return self._save_to_bytes()
    
    # ==================== ДЕНЬ 1 ====================
    
    def _fill_plan_sheet(self):
        """Заполняет лист 'План подбора' (Шаг 3)"""
        submission = self.submissions.get(3)
        if not submission or not submission.structured_data:
            return
        
        data = json.loads(submission.structured_data)
        sheet = self.wb['План подбора']
        start_row = 2
        
        for i, этап in enumerate(data.get('этапы', []), start=1):
            row = start_row + i - 1
            sheet[f'A{row}'] = i
            sheet[f'B{row}'] = этап.get('название', '')
            sheet[f'C{row}'] = этап.get('план', '')
            # D, E остаются пустыми для заполнения руководителем
    
    def _fill_assessment_sheet(self):
        """Заполняет 'ОЦЕНОЧНЫЙ ЛИСТ' (Шаг 6)"""
        submission = self.submissions.get(6)
        if not submission or not submission.structured_data:
            return
        
        data = json.loads(submission.structured_data)
        sheet = self.wb['ОЦЕНОЧНЫЙ ЛИСТ ']
        
        # Soft Skills (строки 5+)
        soft_start = 5
        for i, skill in enumerate(data.get('soft_skills', [])):
            row = soft_start + i
            sheet[f'A{row}'] = skill.get('название', '')
            sheet[f'D{row}'] = skill.get('индикаторы', '')
            sheet[f'E{row}'] = skill.get('вопрос', '')
        
        # Hard Skills (строки 11+)
        hard_start = 11
        for i, skill in enumerate(data.get('hard_skills', [])):
            row = hard_start + i
            sheet[f'A{row}'] = skill.get('название', '')
            sheet[f'D{row}'] = skill.get('индикаторы', '')
            sheet[f'E{row}'] = skill.get('вопрос', '')
        
        # Отсекающие факторы (строки 19+)
        factors_start = 19
        for i, factor in enumerate(data.get('отсекающие_факторы', [])):
            row = factors_start + i
            sheet[f'A{row}'] = factor if isinstance(factor, str) else str(factor)
    
    def _fill_vacancy_sheet(self):
        """Заполняет 'Объявления на Вакансию' (Шаг 12)"""
        submission = self.submissions.get(12)
        if not submission or not submission.structured_data:
            return
        
        data = json.loads(submission.structured_data)
        sheet = self.wb['Объявления на Вакансию']
        
        # Вариант 1: для сайта
        текст_сайт = data.get('для_сайта', {}).get('текст', '') if isinstance(data.get('для_сайта'), dict) else data.get('для_сайта', '')
        sheet['B2'] = текст_сайт
        sheet['C2'] = len(текст_сайт)
        
        # Вариант 2: для мессенджеров
        текст_месс = data.get('для_мессенджеров', {}).get('текст', '') if isinstance(data.get('для_мессенджеров'), dict) else data.get('для_мессенджеров', '')
        sheet['B3'] = текст_месс
        sheet['C3'] = len(текст_месс)
        
        # Вариант 3: для телефона
        текст_тел = data.get('для_телефона', {}).get('текст', '') if isinstance(data.get('для_телефона'), dict) else data.get('для_телефона', '')
        sheet['B4'] = текст_тел
        sheet['C4'] = len(текст_тел)
    
    # ==================== ДЕНЬ 2 ====================
    
    def _fill_active_search_sheet(self):
        """Заполняет 'Карта активного поиска' (Шаг 19)"""
        submission = self.submissions.get(19)
        if not submission or not submission.structured_data:
            return
        
        data = json.loads(submission.structured_data)
        sheet = self.wb['Карта активного поиска']
        
        row = 2
        sheet[f'A{row}'] = "Google поиск"
        sheet[f'B{row}'] = data.get('запрос', '')
        операторы = data.get('использованные_операторы', [])
        sheet[f'C{row}'] = ', '.join(операторы) if isinstance(операторы, list) else str(операторы)
        sheet[f'D{row}'] = data.get('обоснование', '')
    
    def _fill_passive_search_sheet(self):
        """Заполняет 'Карта пассивного поиска' (Шаг 22)"""
        submission = self.submissions.get(22)
        if not submission or not submission.structured_data:
            return
        
        data = json.loads(submission.structured_data)
        sheet = self.wb['Карта пассивного поиска']
        
        row = 2
        for запрос in data.get('запросы', []):
            sheet[f'A{row}'] = запрос.get('соцсеть', '')
            sheet[f'B{row}'] = запрос.get('запрос', '')
            sheet[f'C{row}'] = запрос.get('обоснование', '')
            row += 1
    
    def _fill_market_analysis_sheet(self):
        """Заполняет 'Анализ рынка' (Шаг 25)"""
        submission = self.submissions.get(25)
        if not submission or not submission.structured_data:
            return
        
        data = json.loads(submission.structured_data)
        sheet = self.wb['Анализ рынка']
        
        # Заголовок анализа
        row = 2
        sheet[f'A{row}'] = "АНАЛИЗ РЫНКА"
        sheet[f'A{row}'].font = Font(bold=True, size=12)
        sheet.merge_cells(f'A{row}:F{row}')
        
        row += 2
        
        # Средняя зарплата
        зп = data.get('средняя_зарплата', {})
        if зп:
            sheet[f'A{row}'] = "Средняя зарплата:"
            зп_от = зп.get('от', 0)
            зп_до = зп.get('до', 0)
            медиана = зп.get('медиана', 0)
            sheet[f'B{row}'] = f"{зп_от:,} - {зп_до:,} руб (медиана: {медиана:,})"
            row += 1
        
        # Компании-доноры
        компании = data.get('компании_доноры', [])
        if компании:
            sheet[f'A{row}'] = "Топ компании-доноры:"
            row += 1
            for comp in компании[:3]:
                название = comp.get('название', '')
                кол_резюме = comp.get('резюме', 0)
                sheet[f'B{row}'] = f"• {название} ({кол_резюме} резюме)"
                row += 1
        
        row += 1
        
        # Требования кандидатов
        требования = data.get('требования_кандидатов', [])
        if требования:
            sheet[f'A{row}'] = "Требования кандидатов:"
            row += 1
            for req in требования:
                sheet[f'B{row}'] = f"• {req}"
                row += 1
        
        row += 1
        
        # Дефицитные навыки
        навыки = data.get('дефицитные_навыки', [])
        if навыки:
            sheet[f'A{row}'] = "Дефицитные навыки:"
            row += 1
            for skill in навыки:
                sheet[f'B{row}'] = f"• {skill}"
                row += 1
        
        row += 2
        
        # Рекомендации
        рекомендации = data.get('рекомендации', '')
        if рекомендации:
            sheet[f'A{row}'] = "Рекомендации:"
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            sheet[f'A{row}'] = рекомендации
            sheet.merge_cells(f'A{row}:F{row}')
    
    # ==================== ДЕНЬ 3 ====================
    
    def _fill_speech_modules_sheet(self):
        """Заполняет 'Речевые модули' (Шаги 28-29)"""
        sheet = self.wb['Речевые модули']
        row = 2
        
        # Шаблон недозвона (шаг 28)
        submission_28 = self.submissions.get(28)
        if submission_28 and submission_28.structured_data:
            data_28 = json.loads(submission_28.structured_data)
            sheet[f'A{row}'] = "Шаблон для недозвона"
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            шаблон = data_28.get('шаблон_недозвона', '')
            sheet[f'A{row}'] = шаблон
            sheet.merge_cells(f'A{row}:D{row}')
            row += 2
        
        # Скрипт звонка (шаг 29)
        submission_29 = self.submissions.get(29)
        if submission_29 and submission_29.structured_data:
            data_29 = json.loads(submission_29.structured_data)
            скрипт = data_29.get('скрипт_звонка', {})
            
            sheet[f'A{row}'] = "Скрипт телефонного звонка"
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            этапы = [
                ("1. Приветствие", скрипт.get('приветствие', '')),
                ("2. Цель звонка", скрипт.get('цель', '')),
                ("3. Презентация", скрипт.get('презентация', '')),
                ("4. Вовлечение", скрипт.get('вовлечение', '')),
                ("5. Договоренность", скрипт.get('договоренность', '')),
                ("6. Завершение", скрипт.get('завершение', ''))
            ]
            
            for название, текст in этапы:
                sheet[f'A{row}'] = название
                sheet[f'B{row}'] = текст
                sheet.merge_cells(f'B{row}:D{row}')
                row += 1
    
    def _fill_objections_sheet(self):
        """Заполняет 'Возражения' (Шаг 31)"""
        submission = self.submissions.get(31)
        if not submission or not submission.structured_data:
            return
        
        data = json.loads(submission.structured_data)
        sheet = self.wb['Возражения']
        
        # Заголовки
        sheet['A1'] = "Возражение"
        sheet['B1'] = "Ответ/Отработка"
        sheet['A1'].font = Font(bold=True)
        sheet['B1'].font = Font(bold=True)
        
        row = 2
        for возр in data.get('возражения', []):
            sheet[f'A{row}'] = возр.get('возражение', '')
            sheet[f'B{row}'] = возр.get('ответ', '')
            row += 1
    
    # ==================== ОТЧЕТ ====================
    
    def _add_onboarding_report_sheet(self):
        """Создает новый лист 'Отчет по онбордингу' для всех 36 шагов"""
        if 'Отчет онбординга' in self.wb.sheetnames:
            del self.wb['Отчет онбординга']
        
        report_sheet = self.wb.create_sheet('Отчет онбординга', 0)  # Первый лист
        
        # Заголовок
        report_sheet['A1'] = 'ОТЧЕТ ПО ОНБОРДИНГУ (3 ДНЯ)'
        report_sheet['A1'].font = Font(size=16, bold=True)
        report_sheet.merge_cells('A1:F1')
        
        # Метаданные
        first_submission = next(iter(self.submissions.values()), None)
        if first_submission and hasattr(first_submission, 'user'):
            user = first_submission.user
            report_sheet['A3'] = 'Стажер:'
            report_sheet['B3'] = user.full_name if hasattr(user, 'full_name') else 'N/A'
            report_sheet['A4'] = 'Telegram:'
            report_sheet['B4'] = f'@{user.username}' if hasattr(user, 'username') else 'N/A'
        
        report_sheet['A5'] = 'Дата завершения:'
        report_sheet['B5'] = datetime.now().strftime('%d.%m.%Y')
        
        # Таблица результатов
        report_sheet['A7'] = 'Шаг'
        report_sheet['B7'] = 'Название'
        report_sheet['C7'] = 'День'
        report_sheet['D7'] = 'Оценка'
        report_sheet['E7'] = 'Время'
        report_sheet['F7'] = 'Статус'
        
        for col in ['A7', 'B7', 'C7', 'D7', 'E7', 'F7']:
            report_sheet[col].font = Font(bold=True)
        
        row = 8
        for step_id in range(1, 37):  # 36 шагов
            submission = self.submissions.get(step_id)
            if submission:
                день = "День 1" if step_id <= 13 else "День 2" if step_id <= 26 else "День 3"
                
                report_sheet[f'A{row}'] = step_id
                report_sheet[f'B{row}'] = submission.step.title if hasattr(submission, 'step') else 'N/A'
                report_sheet[f'C{row}'] = день
                
                # Оценка
                score = submission.evaluation_score if hasattr(submission, 'evaluation_score') else None
                report_sheet[f'D{row}'] = score if score is not None else '-'
                
                # Время
                if hasattr(submission, 'get_completion_time_minutes'):
                    time_min = submission.get_completion_time_minutes()
                    report_sheet[f'E{row}'] = f'{time_min:.0f} мин' if time_min else '-'
                else:
                    report_sheet[f'E{row}'] = '-'
                
                # Статус
                status = submission.status if hasattr(submission, 'status') else 'pending'
                report_sheet[f'F{row}'] = '✅' if status == 'approved' else '⏳'
                row += 1
        
        # Сводка по дням
        row += 2
        report_sheet[f'A{row}'] = 'СВОДКА ПО ДНЯМ'
        report_sheet[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        день1_оценка = self._calculate_avg_score(range(1, 14))
        день2_оценка = self._calculate_avg_score(range(14, 27))
        день3_оценка = self._calculate_avg_score(range(27, 37))
        
        report_sheet[f'A{row}'] = 'День 1 (Основы подбора):'
        report_sheet[f'B{row}'] = f'{день1_оценка:.2f} / 5' if день1_оценка > 0 else 'N/A'
        row += 1
        report_sheet[f'A{row}'] = 'День 2 (Сорсинг и анализ):'
        report_sheet[f'B{row}'] = f'{день2_оценка:.2f} / 5' if день2_оценка > 0 else 'N/A'
        row += 1
        report_sheet[f'A{row}'] = 'День 3 (Звонки и интервью):'
        report_sheet[f'B{row}'] = f'{день3_оценка:.2f} / 5' if день3_оценка > 0 else 'N/A'
        
        row += 2
        
        # Сводка по компетенциям
        report_sheet[f'A{row}'] = 'СВОДКА ПО КОМПЕТЕНЦИЯМ'
        report_sheet[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        теория = self._calculate_avg_score([2, 5, 7, 8, 14, 16, 20, 23, 27, 30, 33, 34])
        практика = self._calculate_avg_score([3, 6, 12, 19, 22, 25, 28, 29, 31, 35])
        анализ = self._calculate_avg_score([4, 9, 10, 11, 15, 17, 18, 21, 24])
        
        report_sheet[f'A{row}'] = 'Теоретические знания:'
        report_sheet[f'B{row}'] = f'{теория:.2f} / 5' if теория > 0 else 'N/A'
        row += 1
        report_sheet[f'A{row}'] = 'Практические навыки:'
        report_sheet[f'B{row}'] = f'{практика:.2f} / 5' if практика > 0 else 'N/A'
        row += 1
        report_sheet[f'A{row}'] = 'Аналитическое мышление:'
        report_sheet[f'B{row}'] = f'{анализ:.2f} / 5' if анализ > 0 else 'N/A'
        
        # Комментарии наставников (шаги 13, 26, 36)
        row += 2
        report_sheet[f'A{row}'] = 'КОММЕНТАРИИ НАСТАВНИКА'
        report_sheet[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for день, step_id in [("День 1", 13), ("День 2", 26), ("День 3", 36)]:
            step = self.submissions.get(step_id)
            if step and hasattr(step, 'text_answer') and step.text_answer:
                report_sheet[f'A{row}'] = f'{день}:'
                report_sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                report_sheet[f'A{row}'] = step.text_answer
                report_sheet.merge_cells(f'A{row}:F{row}')
                row += 2
    
    def _calculate_avg_score(self, step_ids) -> float:
        """Вычисляет среднюю оценку по списку шагов"""
        scores = []
        for step_id in step_ids:
            submission = self.submissions.get(step_id)
            if submission and hasattr(submission, 'evaluation_score') and submission.evaluation_score:
                scores.append(submission.evaluation_score)
        return sum(scores) / len(scores) if scores else 0.0
    
    def _save_to_bytes(self) -> bytes:
        """Сохраняет workbook в bytes"""
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
