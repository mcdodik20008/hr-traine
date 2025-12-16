"""
Новый генератор Excel отчета для онбординга
4 листа: Summary + День 1, 2, 3 с LLM-оценками
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import List, Dict, Optional
from io import BytesIO
import re


class SimpleOnboardingReportGenerator:
    """Генератор отчета с LLM-оценками для каждого дня"""
    
    def __init__(self, user_submissions: List):
        """
        Args:
            user_submissions: Список OnboardingSubmission для пользователя
        """
        # Filter out submissions where step is None to prevent AttributeError
        self.submissions = {s.step.order: s for s in user_submissions if s.step is not None}
        self.wb = Workbook()
        self.user = None
        if user_submissions and hasattr(user_submissions[0], 'user'):
            self.user = user_submissions[0].user
        
        # Кэш для LLM оценок
        self.llm_evaluations = {}
        
    async def generate_async(self) -> bytes:
        """Асинхронная генерация Excel отчета с LLM оценками"""
        # Удаляем стандартный лист
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Сначала оцениваем все ответы через LLM
        await self._evaluate_all_answers()
        
        # Создаем листы отчета
        self._create_summary_sheet()
        await self._create_day_sheet(1, 1, 13)
        await self._create_day_sheet(2, 14, 26)
        await self._create_day_sheet(3, 27, 36)
        
        return self._save_to_bytes()
    
    def generate(self) -> bytes:
        """Синхронная версия для обратной совместимости (без LLM оценок)"""
        # Удаляем стандартный лист
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Создаем листы без LLM оценок
        self._create_summary_sheet()
        self._create_day_sheet_sync(1, 1, 13)
        self._create_day_sheet_sync(2, 14, 26)
        self._create_day_sheet_sync(3, 27, 36)
        
        return self._save_to_bytes()
    
    async def _evaluate_all_answers(self):
        """Оцениваем все текстовые ответы через LLM"""
        from app.core.llm_client import llm_client
        import logging
        
        logger = logging.getLogger(__name__)
        logger.info(f"Starting LLM evaluation for {len(self.submissions)} submissions")
        
        for step_order, submission in self.submissions.items():
            text_answer = getattr(submission, 'text_answer', None)
            
            logger.info(f"Step {step_order}: text_answer length = {len(text_answer) if text_answer else 0}")
            
            # Пропускаем пустые ответы и простые подтверждения
            if not text_answer or text_answer == 'Completed':
                logger.info(f"Step {step_order}: Skipped (empty or 'Completed')")
                continue
            
            # Убираем слишком строгую проверку длины - оцениваем все текстовые ответы
            if len(text_answer.strip()) < 3:
                logger.info(f"Step {step_order}: Skipped (too short: {len(text_answer.strip())} chars)")
                continue
            
            step = submission.step
            step_description = getattr(step, 'description', '')
            step_title = getattr(step, 'title', f'Шаг {step_order}')
            
            logger.info(f"Step {step_order}: Evaluating with LLM. Answer length: {len(text_answer)}")
            
            # Создаем промпт для оценки
            prompt = f"""Ты наставник HR-стажёра. Оцени ответ стажёра на задание онбординга.

Задание (Шаг {step_order}: {step_title}):
{step_description}

Ответ стажёра:
\"\"\"{text_answer}\"\"\"

Оцени ответ по шкале от 1 до 10 и дай краткий отзыв.

Формат ответа:
Оценка: [число от 1 до 10]
Отзыв: [2-3 предложения с конструктивной обратной связью]
Сильные стороны: [что хорошо]
Что улучшить: [конкретные рекомендации]
"""
            
            try:
                feedback = await llm_client.generate_response(prompt)
                logger.info(f"Step {step_order}: LLM response received: {feedback[:100]}...")
                
                # Парсим ответ LLM
                score = self._extract_score(feedback)
                logger.info(f"Step {step_order}: Extracted score: {score}")
                
                self.llm_evaluations[step_order] = {
                    'score': score,
                    'feedback': feedback,
                    'raw_text': text_answer
                }
            except Exception as e:
                logger.error(f"Step {step_order}: LLM error: {str(e)}", exc_info=True)
                # Если LLM недоступен, ставим средний балл
                self.llm_evaluations[step_order] = {
                    'score': 5.0,
                    'feedback': f'Оценка недоступна (ошибка LLM: {str(e)})',
                    'raw_text': text_answer
                }
        
        logger.info(f"LLM evaluation completed. Total evaluations: {len(self.llm_evaluations)}")
    
    def _extract_score(self, feedback: str) -> float:
        """Извлекает оценку из ответа LLM"""
        # Ищем "Оценка: X" или просто число от 1 до 10
        match = re.search(r'Оценка:\s*([0-9]+(?:[.,][0-9]+)?)', feedback, re.IGNORECASE)
        if match:
            score_str = match.group(1).replace(',', '.')
            score = float(score_str)
            return max(1.0, min(10.0, score))
        
        # Альтернативный поиск просто числа
        match = re.search(r'\b([1-9]|10)(?:[.,]([0-9]))?\b', feedback)
        if match:
            score = float(match.group(0).replace(',', '.'))
            return max(1.0, min(10.0, score))
        
        return 5.0  # По умолчанию средний балл
    
    def _calculate_overall_score(self) -> float:
        """Вычисляет общую оценку (среднее всех LLM оценок)"""
        if not self.llm_evaluations:
            return 0.0
        
        scores = [eval_data['score'] for eval_data in self.llm_evaluations.values()]
        return sum(scores) / len(scores) if scores else 0.0
    
    def _generate_overview(self, overall_score: float) -> str:
        """Генерирует краткий обзор на основе оценок"""
        if not self.llm_evaluations:
            return "Недостаточно данных для формирования обзора."
        
        # Группируем по дням
        day1_scores = [ev['score'] for order, ev in self.llm_evaluations.items() if 1 <= order <= 13]
        day2_scores = [ev['score'] for order, ev in self.llm_evaluations.items() if 14 <= order <= 26]
        day3_scores = [ev['score'] for order, ev in self.llm_evaluations.items() if 27 <= order <= 36]
        
        avg_day1 = sum(day1_scores) / len(day1_scores) if day1_scores else 0
        avg_day2 = sum(day2_scores) / len(day2_scores) if day2_scores else 0
        avg_day3 = sum(day3_scores) / len(day3_scores) if day3_scores else 0
        
        # Формируем обзор
        overview_parts = []
        
        # Общая оценка
        if overall_score >= 8.0:
            overview_parts.append("🌟 Отличная работа! Стажёр показал высокий уровень понимания материала.")
        elif overall_score >= 6.0:
            overview_parts.append("✅ Хорошая работа. Стажёр справился с большинством заданий на достойном уровне.")
        elif overall_score >= 4.0:
            overview_parts.append("⚠️ Удовлетворительно. Есть понимание основ, но требуется дополнительная проработка.")
        else:
            overview_parts.append("❌ Требуется значительное улучшение. Рекомендуется повторное прохождение.")
        
        # Анализ по дням
        best_day = max([(avg_day1, "День 1"), (avg_day2, "День 2"), (avg_day3, "День 3")], key=lambda x: x[0])
        worst_day = min([(avg_day1, "День 1"), (avg_day2, "День 2"), (avg_day3, "День 3")], key=lambda x: x[0])
        
        if best_day[0] > 0:
            overview_parts.append(f"\n🎯 Сильная сторона: {best_day[1]} (средний балл {best_day[0]:.1f}/10)")
        
        if worst_day[0] > 0 and worst_day[0] < 6.0:
            overview_parts.append(f"📌 Требует внимания: {worst_day[1]} (средний балл {worst_day[0]:.1f}/10)")
        
        return "\n".join(overview_parts)
    
    def _create_summary_sheet(self):
        """Создает лист с общей сводкой и оценкой 1-10"""
        sheet = self.wb.create_sheet('📊 Summary', 0)
        
        # Заголовок
        sheet['A1'] = 'ОТЧЕТ ПО ОНБОРДИНГУ HR TRAINEE'
        sheet['A1'].font = Font(size=18, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        sheet.merge_cells('A1:D1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.row_dimensions[1].height = 30
        
        # Информация о стажёре
        row = 3
        if self.user:
            sheet[f'A{row}'] = '👤 Стажёр:'
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'] = getattr(self.user, 'full_name', 'N/A')
            sheet.merge_cells(f'B{row}:D{row}')
            row += 1
            
            sheet[f'A{row}'] = '📱 Telegram:'
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'] = f"@{getattr(self.user, 'username', 'N/A')}"
            sheet.merge_cells(f'B{row}:D{row}')
            row += 1
        
        sheet[f'A{row}'] = '📅 Дата:'
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'B{row}'] = datetime.now().strftime('%d.%m.%Y %H:%M')
        sheet.merge_cells(f'B{row}:D{row}')
        row += 2
        
        # Общая оценка
        overall_score = self._calculate_overall_score()
        
        sheet[f'A{row}'] = '🎯 ОБЩАЯ ОЦЕНКА'
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        sheet.merge_cells(f'A{row}:D{row}')
        row += 1
        
        sheet[f'A{row}'] = f'{overall_score:.1f} / 10'
        sheet[f'A{row}'].font = Font(size=36, bold=True, color='4472C4')
        sheet.merge_cells(f'A{row}:D{row}')
        sheet[f'A{row}'].alignment = Alignment(horizontal='center')
        sheet.row_dimensions[row].height = 50
        row += 2
        
        # Краткий обзор
        sheet[f'A{row}'] = '📝 КРАТКИЙ ОБЗОР'
        sheet[f'A{row}'].font = Font(size=12, bold=True)
        sheet.merge_cells(f'A{row}:D{row}')
        row += 1
        
        overview = self._generate_overview(overall_score)
        sheet[f'A{row}'] = overview
        sheet.merge_cells(f'A{row}:D{row}')
        sheet[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        sheet.row_dimensions[row].height = 80
        row += 2
        
        # Статистика
        sheet[f'A{row}'] = '📈 СТАТИСТИКА'
        sheet[f'A{row}'].font = Font(size=12, bold=True)
        sheet.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Таблица статистики
        stats_data = [
            ('Всего шагов:', len(self.submissions)),
            ('Оценено LLM:', len(self.llm_evaluations)),
        ]
        
        # Средние оценки по дням
        day1_scores = [ev['score'] for order, ev in self.llm_evaluations.items() if 1 <= order <= 13]
        day2_scores = [ev['score'] for order, ev in self.llm_evaluations.items() if 14 <= order <= 26]
        day3_scores = [ev['score'] for order, ev in self.llm_evaluations.items() if 27 <= order <= 36]
        
        if day1_scores:
            stats_data.append(('День 1 (среднее):', f"{sum(day1_scores)/len(day1_scores):.1f}/10"))
        if day2_scores:
            stats_data.append(('День 2 (среднее):', f"{sum(day2_scores)/len(day2_scores):.1f}/10"))
        if day3_scores:
            stats_data.append(('День 3 (среднее):', f"{sum(day3_scores)/len(day3_scores):.1f}/10"))
        
        for label, value in stats_data:
            sheet[f'A{row}'] = label
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'] = value
            row += 1
        
        # Настройка ширины столбцов
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
    
    async def _create_day_sheet(self, day_num: int, start_step: int, end_step: int):
        """Создает лист с оценками для конкретного дня (асинхронная версия)"""
        self._create_day_sheet_impl(day_num, start_step, end_step)
    
    def _create_day_sheet_sync(self, day_num: int, start_step: int, end_step: int):
        """Создает лист с оценками для конкретного дня (синхронная версия)"""
        self._create_day_sheet_impl(day_num, start_step, end_step)
    
    def _create_day_sheet_impl(self, day_num: int, start_step: int, end_step: int):
        """Реализация создания листа для дня"""
        sheet = self.wb.create_sheet(f'📅 День {day_num}')
        
        # Заголовок
        sheet['A1'] = f'ДЕНЬ {day_num} - ДЕТАЛЬНАЯ ОЦЕНКА'
        sheet['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        sheet.merge_cells('A1:D1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.row_dimensions[1].height = 25
        
        row = 3
        
        # Проходим по всем шагам дня
        for step_order in range(start_step, end_step + 1):
            submission = self.submissions.get(step_order)
            
            if not submission:
                continue
            
            step = submission.step
            step_title = getattr(step, 'title', f'Шаг {step_order}')
            step_description = getattr(step, 'description', '')
            text_answer = getattr(submission, 'text_answer', None)
            
            # Заголовок шага
            sheet[f'A{row}'] = f'Шаг {step_order}: {step_title}'
            sheet[f'A{row}'].font = Font(size=11, bold=True, color='FFFFFF')
            sheet[f'A{row}'].fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
            sheet.merge_cells(f'A{row}:D{row}')
            sheet.row_dimensions[row].height = 20
            row += 1
            
            # Описание задания (если есть текстовый ответ)
            if text_answer and text_answer != 'Completed':
                if step_description:
                    sheet[f'A{row}'] = '📋 Задание:'
                    sheet[f'A{row}'].font = Font(bold=True, size=10)
                    sheet[f'B{row}'] = step_description
                    sheet.merge_cells(f'B{row}:D{row}')
                    sheet[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                    row += 1
                
                # Ответ стажёра
                sheet[f'A{row}'] = '✍️ Ответ:'
                sheet[f'A{row}'].font = Font(bold=True, size=10)
                sheet[f'B{row}'] = text_answer
                sheet.merge_cells(f'B{row}:D{row}')
                sheet[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                row += 1
                
                # LLM оценка
                if step_order in self.llm_evaluations:
                    eval_data = self.llm_evaluations[step_order]
                    score = eval_data['score']
                    feedback = eval_data['feedback']
                    
                    # Оценка
                    sheet[f'A{row}'] = '⭐ Оценка:'
                    sheet[f'A{row}'].font = Font(bold=True, size=10)
                    sheet[f'B{row}'] = f"{score:.1f} / 10"
                    sheet[f'B{row}'].font = Font(size=11, bold=True, color='C00000' if score < 5 else '375623')
                    row += 1
                    
                    # Фидбек от LLM
                    sheet[f'A{row}'] = '💬 Фидбек:'
                    sheet[f'A{row}'].font = Font(bold=True, size=10)
                    sheet[f'B{row}'] = feedback
                    sheet.merge_cells(f'B{row}:D{row}')
                    sheet[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                    row += 1
            else:
                # Если нет текстового ответа, показываем базовую информацию
                sheet[f'A{row}'] = '✅ Статус:'
                sheet[f'A{row}'].font = Font(bold=True, size=10)
                status = getattr(submission, 'status', 'pending')
                sheet[f'B{row}'] = 'Выполнено' if status in ['checked', 'approved'] else 'В процессе'
                row += 1
            
            # Пустая строка между шагами
            row += 1
        
        # Настройка ширины столбцов
        sheet.column_dimensions['A'].width = 18
        sheet.column_dimensions['B'].width = 60
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
    
    def _save_to_bytes(self) -> bytes:
        """Сохраняет workbook в bytes"""
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
