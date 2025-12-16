import openpyxl
import pandas as pd

# Открываем файл
wb = openpyxl.load_workbook('Карта_поиска_для_проектной_работы.xlsx', data_only=True)

print("="*80)
print("СТРУКТУРА ФАЙЛА 'Карта поиска'")
print("="*80)

for sheet in wb.worksheets:
    print(f"\n📋 Лист: '{sheet.title}'")
    print(f"   Размер: {sheet.max_row} строк x {sheet.max_column} колонок")
    
    # Показываем заголовки (первая непустая строка)
    headers = []
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        if any(cell for cell in row):
            non_empty = [cell for cell in row if cell is not None]
            if non_empty:
                headers = non_empty
                break
    
    if headers:
        print(f"   Заголовки: {headers[:6]}")  # Первые 6

# Детально смотрим ОЦЕНОЧНЫЙ ЛИСТ
print("\n" + "="*80)
print("ДЕТАЛЬНЫЙ АНАЛИЗ: ОЦЕНОЧНЫЙ ЛИСТ")
print("="*80)

sheet = wb['ОЦЕНОЧНЫЙ ЛИСТ ']
for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), 1):
    cells = [str(c) if c is not None else '' for c in row[:8]]
    if any(cells):
        print(f"{row_idx:2d}| {' | '.join(cells)}")

# План подбора
print("\n" + "="*80)
print("ДЕТАЛЬНЫЙ АНАЛИЗ: ПЛАН ПОДБОРА")
print("="*80)

sheet = wb['План подбора']
for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
    cells = [str(c) if c is not None else '' for c in row[:6]]
    if any(cells):
        print(f"{row_idx:2d}| {' | '.join(cells)}")

# Объявления
print("\n" + "="*80)
print("ДЕТАЛЬНЫЙ АНАЛИЗ: ОБЪЯВЛЕНИЯ НА ВАКАНСИЮ")
print("="*80)

sheet = wb['Объявления на Вакансию']
for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
    cells = [str(c)[:40] if c is not None else '' for c in row[:5]]
    if any(cells):
        print(f"{row_idx:2d}| {' | '.join(cells)}")
